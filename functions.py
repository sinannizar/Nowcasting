from sklearn.linear_model import LogisticRegression
from numpy import genfromtxt
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image
from mpl_toolkits.basemap import Basemap
from scipy import interpolate
from shapely.geometry import Polygon
import geopandas as gpd
from openpyxl import load_workbook
from pyproj import CRS


def Logistic(d15):
    # Read the training data
    train = genfromtxt('Data/TrainingData.csv', delimiter=',')

    # Read the d15/d70 value
    X_test = [d15]

    # Read user nowcasting parameters from the user
    print("Enter the value of rb: ")
    X_test.append(float(input()))
    print("Enter the value of Tb: ")
    X_test.append(float(input()))
    print("Enter the value of T14: ")
    X_test.append(float(input()))
    print("Enter the value of TL: ")
    X_test.append(float(input()))
    print("Enter the value of Dz: ")
    X_test.append(float(input()))
    X_test = np.array([X_test])

    X_train = train[:, :6]
    y_train = train[:, -1]

    # Create a logistic regression object
    logreg = LogisticRegression()

    # Train the logistic regression model on the training data
    logreg.fit(X_train, y_train)

    # Predict the classes of the testing data
    y_pred = logreg.predict(X_test)

    if y_pred[0] == 0:
        prob = format(logreg.predict_proba(X_test)[0][0], '.1f')
        print("A Normal Rainfall Event is expected with a probability of %s" % prob)
    else:
        prob = format(logreg.predict_proba(X_test)[0][1], '.1f')
        print("A Heavy/Extreme Rainfall Event is expected with a probability of %s" % prob)


def norm(band, band_min, band_max):
    x = (band - band_min)/(band_max - band_min)
    x[x > 1] = 1
    x[x < 0] = 0
    return x


def fill(array):
    x = np.arange(0, array.shape[1])
    y = np.arange(0, array.shape[0])
    #mask invalid values
    array = np.ma.masked_invalid(array)
    xx, yy = np.meshgrid(x, y)
    #get only the valid values
    x1 = xx[~array.mask]
    y1 = yy[~array.mask]
    newarr = array[~array.mask]

    GD1 = interpolate.griddata((x1, y1), newarr.ravel(), (xx, yy), method='nearest')
    return GD1


def GetShape():
    ltt = []
    lgg = []
    print("Enter the coordinates")
    for i in range(4):
        ltt.append(float(input()))
        lgg.append(float(input()))
    ltt.append(ltt[0])
    lgg.append(lgg[0])
    print("Creating polygon")
    polygon_geom = Polygon(zip(lgg, ltt))
    crs = CRS('EPSG:4326')
    polygon = gpd.GeoDataFrame(index=[0], crs=crs, geometry=[polygon_geom])
    polygon.to_file(filename='Data/Shapefile/polygon.shp', driver="ESRI Shapefile")


def count(cer_val):
    d15 = 0
    d70 = 0
    for i in cer_val:
        if 10 < i <= 15:
            d15 += 1
        if 65 < i <= 70:
            d70 += 1
    return d15/d70


def RGBmap(data):
    print("Computing RGB composite")

    lat = np.array(data['latitude'][:])
    lon = np.array(data['longitude'][:])

    red = np.array(data['reflectance_vis'][:])
    green = np.array(data['temperature_sir'][:])
    blue = np.array(data['temperature_ir'][:])

    red = norm(red.astype(float), 0, 1)
    green = norm(green.astype(float), 203, 323)
    blue = norm(blue.astype(float), 203, 323)
    im = Image.open('Data/logo.png')

    y = fill(lat)
    x = fill(lon)

    rgb = np.dstack((red, green, blue))

    fig, ax = plt.subplots(figsize=(10, 10))
    ax.tick_params(axis="y", direction="out", length=8, width=1.5)
    ax.tick_params(axis="x", direction="out", length=8, width=1.5)

    for axis in ['top', 'bottom', 'left', 'right']:
        ax.spines[axis].set_linewidth(2)
    m = Basemap()
    m.readshapefile("Data/Shapefile/STATE", 'coastline')

    mesh_rgb = rgb[:, :-1, :]
    colorTuple = mesh_rgb.reshape((mesh_rgb.shape[0] * mesh_rgb.shape[1]), 3)
    colorTuple = np.insert(colorTuple, 3, 1.0, axis=1)

    m.pcolormesh(x, y, rgb[:, :, 1], color=colorTuple)
    plt.grid(linestyle='dashed', linewidth=0.5)
    plt.yticks(np.arange(0, 22, step=1))
    plt.xticks(np.arange(65, 87, step=1))
    plt.xlim(65, 85)
    plt.ylim(0, 20)
    plt.margins(1)
    plt.subplots_adjust(top=1)
    plt.text(66.5, 20.3, r"$\bf{" + 'KSCSTE' + "}$" + '\nInstitute for Climate Change Studies', fontsize=14)
    plt.text(66, -1.4, 'Cloud RGB composite image inherited from Rosenfeld and Lensky (1998).', fontsize=12,
             weight='bold')
    plt.text(64.5, -3.5,
             r"$\bf{" + 'Color ' + "}$" + ' ' + r"$\bf{" + 'scheme:' + "}$" +
             ' Water  clouds  that do not precipitate appear  white, whereas large drops that\nare typical  to  '
             'precipitating  clouds  appear  pink.  Supercooled water  clouds appear as yellow.\nCold and thick '
             'clouds with large ice particles at top appear red. Optically thick cloud with small\nice particles '
             'near their tops appear orange.',
             fontsize=12)
    fig.figimage(im, 85, 1410)

    print("Saving RGB image")
    plt.savefig('Output/OutputRGB.jpg', bbox_inches='tight', pad_inches=0.2, dpi=150)


def mask(data):
    lat = np.array(data['latitude'][:])
    lon = np.array(data['longitude'][:])
    CER = np.array(data['cloud_particle_size'][:])
    CTT = np.array(data['cloud_top_temperature'][:])
    print("Masking to shapefile")
    GetShape()
    shapefile = "Data/Shapefile/polygon.shp"
    box = gpd.read_file(shapefile)
    points = gpd.read_file("Data/Shapefile/IndiaPoints.shp")
    MaskPoints = gpd.clip(points, box)
    mask = np.empty([875, 1700])
    for pt in MaskPoints.index:
        mask[MaskPoints['Field3'][pt]][MaskPoints['Field4'][pt]] = 1
    y = fill(lat)
    x = fill(lon)
    CER = np.multiply(mask, CER)
    CTT = np.multiply(mask, CTT)
    CER[CER == 0] = np.nan
    CTT[CTT == 0] = np.nan
    wt = load_workbook('Data/CTTvCER_Mother.xlsx')
    sheet1 = wt['Sheet1']
    k = 2
    print("Saving to excel")
    cer = []
    for i in range(CER.shape[0]):
        for j in range(CER.shape[1]):
            if not np.isnan(CER[i][j]) and not np.isnan(CTT[i][j]):
                sheet1.cell(row=k, column=1).value = CER[i][j]
                sheet1.cell(row=k, column=2).value = CTT[i][j]
                cer.append(CER[i][j])
                k += 1
    d15_70 = count(np.array(cer))
    wt.save('Output\\CTTvCER.xlsx')
    wt.close()
    print("Process Completed")
    print("--------------------------------------------------------------------")
    return d15_70

